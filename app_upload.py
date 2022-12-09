"""This module contains functions used to upload data to the app."""

import pandas as pd
import numpy as np
import streamlit as st
from openpyxl import load_workbook
import itertools
from streamlit_folium import folium_static
import folium
import os
import json
from os import listdir, replace
from os.path import isfile, join
import traceback
import configparser

from scenario import read_scenario, initialize_scenario


@st.cache(suppress_st_warning=True, show_spinner=False, allow_output_mutation=True)
def locate_warehouse(ref_name):
    """Map a city to its respective warehouse location. 
    Args:
        ref_name (str): a city in Zambia
    Returns:
            (str): a warehouse location
            
            (str): returns an empty string    
    """

    country_config = configparser.ConfigParser()
    country_config.read("./country_cfg.toml")
    country_name  = "COUNTRY" if country_config.get("country", "country_name").title() =="" else country_config.get("country", "country_name").title()
    with open(f'./data/{country_name}/warehouse_mapping.json') as f:
        REF_WAREHOUSE_MAPPING = json.load(f) 
    for k in REF_WAREHOUSE_MAPPING: 
        if k.upper() in ref_name.upper(): 
            return REF_WAREHOUSE_MAPPING[k]
    return ""

def app_upload(session_state):
    """Get all files in data directory and process to include only correct format.
    Args:
        session_state (object): a new SessionState object 
    """     

    st.markdown("## Upload File 📤", unsafe_allow_html=True)
    st.markdown("***")

    # get all files in data directory and process to include only correct format
    data_path = f'./data/{session_state.country.title()}/'
    only_files = [f for f in listdir(data_path) if isfile(join(data_path, f))
                    and (len(f.split(' ')) >= 2) and f.split(' ')[1] == 'DRT']

    country_files = [f for f in only_files if f.split(' ')[0].title() == session_state.country.title()]
    if len(country_files) != 0:
        ref_index = only_files.index(country_files[0])
    else: 
        only_files = [""] + only_files
        ref_index = 0

    ref_file = st.selectbox(
        label='Select reference file',
        options=only_files,
        index=ref_index,
        format_func=(lambda x : x.replace('.xlsx', ''))
    )

    if ref_file == "": 
        st.stop()

    session_state.ref_file = join(data_path, ref_file)
    session_state.warehouse = locate_warehouse(session_state.ref_file)
    if session_state.warehouse == "": 
        st.error("The reference file has yet to be mapped to a dispatch origin. Please maintain REF_WAREHOUSE_MAPPING in code")
        st.stop()
    st.markdown(f"Dispatch Origin: {session_state.warehouse}")

    """ Define spacing """
    left_space, upload_title, right_space = st.columns([0.5, 2, 0.5])
    left_space, warehouse_expert_info, space, scenario_info, right_space = st.columns([0.5, 0.7, 0.2, 0.7, 0.5])
    left_space, uploader_space, right_space = st.columns([0.6, 2, 1])
    left_space, upload_button_space = st.columns([2, 1])

    """ Define text for user to read """
    upload_title_text = "### Please upload either an Order Evaluation file or a Scenario file"
    raw_info_text = "#### Upload an order evaluation file to: \n - create a new scenario \n - peform a new optimization"
    scenario_info_text = "#### Upload a scenario file to: \n - view previous results  \n - update and optimize"
    upload_title.markdown(upload_title_text)
    warehouse_expert_info.markdown(raw_info_text)
    scenario_info.markdown(scenario_info_text)

    """ Create uploader object """
    uploaded_file = uploader_space.file_uploader("", type=["xlsx", "xlsm"], key="OptFile")
    if uploaded_file is not None:
        upload = upload_button_space.button("Upload")
        if upload:
            session_state.scenario_data = None
            session_state.scenario_file = ""
            session_state.user_file = os.path.join(".", "filestore", uploaded_file.name)
            book = load_workbook(uploaded_file)
            if "Metadata" in book.sheetnames:
                """ If scenario file """
                """ Save file to filestore folder"""
                with open(session_state.user_file, 'wb') as f:
                    f.write(uploaded_file.getbuffer())
                session_state.scenario_data = read_scenario(session_state.user_file)
                session_state.scenario_data["Facility_DF"]['type'] = session_state.scenario_data["Facility_DF"].apply(lambda f: 'Warehouse' if f['facility']==session_state.warehouse else f['type'], axis=1)
                session_state.scenario_file = os.path.join(".", "filestore", uploaded_file.name)
                st.experimental_rerun()
            else: 
                with open(session_state.user_file, 'wb') as f:
                    f.write(uploaded_file.getbuffer())
                if "Delivery" not in book.sheetnames:
                    validate, message = validate_order_data(session_state, uploaded_file)
                    if validate:
                        uploader_space.success(message)
                    else:
                        uploader_space.error(message)
                        st.stop()

    st.markdown("***")

    if session_state.scenario_data is not None:
        display_scenario_data(session_state)
    else: 
        if session_state.user_file != "" and uploaded_file is not None:
            display_raw_data(session_state)

def app_upload_batch(session_state, uploaded_file):
    """Get all files in data directory and process to include only correct format for batch processing
    Args:
        session_state (object): a new SessionState object 
        uploaded_file (object): file object from streamlit
    Returns:
        (bool): True if successful
    """     

    session_state.scenario_data = None
    session_state.scenario_file = ""
    session_state.user_file = os.path.join(".", "filestore", uploaded_file.name)
    book = load_workbook(uploaded_file)
    if "Metadata" in book.sheetnames:
        """ If scenario file """
        """ Save file to filestore folder"""
        with open(session_state.user_file, 'wb') as f:
            f.write(uploaded_file.getbuffer())
        session_state.scenario_data = read_scenario(session_state.user_file)
        session_state.scenario_data["Facility_DF"]['type'] = session_state.scenario_data["Facility_DF"].apply(lambda f: 'Warehouse' if f['facility']==session_state.warehouse else f['type'], axis=1)
        session_state.scenario_file = os.path.join(".", "filestore", uploaded_file.name)
    else: 
        with open(session_state.user_file, 'wb') as f:
            f.write(uploaded_file.getbuffer())
        if "Delivery" not in book.sheetnames:
            validate, message = validate_order_data(session_state, uploaded_file)
            if not validate:
                return False

    ref_facility = pd.read_excel(session_state.ref_file, sheet_name = "Facility")  
    ref_facility["facility_id"] = ref_facility["facility_id"].apply(lambda c: str(c)[:8])
    ref_facility["latitude"] = ref_facility["latitude"].astype(float)
    ref_facility["longitude"] = ref_facility["longitude"].astype(float)
    ref_facility['type'] = ref_facility.apply(lambda f: 'Warehouse' if f['facility']==session_state.warehouse else f['type'], axis=1)

    if "Delivery" in pd.ExcelFile(session_state.user_file).sheet_names:
        deliveries = pd.read_excel(session_state.user_file, sheet_name='Delivery')
        order_data = deliveries.groupby(['Dispatch Destination']).agg({'Loading Volume':'sum'}).reset_index()
        order_data.columns = ['facility', 'vol']
        
        dispatch_locations = order_data['facility'].to_list()
        ref_locations = ref_facility['facility'].to_list()
        missed_facilities = [fac for fac in dispatch_locations if fac not in ref_locations]
        if len(missed_facilities) > 0: 
            st.markdown(f"#### Warning: The following dispatch destinations are not recognized. Please review and maintain reference data as-needed: \n{', '.join(missed_facilities)}")

        order_data = ref_facility.merge(order_data, on = "facility", how="left")
        #st.table(order_data)
    else: 
        order_data = pd.read_excel(session_state.user_file).sort_values("Facility")
        order_data["Facility ID"] = order_data["Facility ID"].apply(lambda c: str(c)[:8])
        order_data = ref_facility.merge(order_data, right_on = "Facility ID", left_on = "facility_id", how="left").rename(columns = {"Volume (cubic meters)" : "vol"})

    order_keys = order_data[order_data["vol"] > 0].facility
    order_data_dict = {key: True for key in order_keys}
  
    filtered_facilities = list(itertools.compress(order_data_dict.keys(), order_data_dict.values()))

    order_data = order_data[order_data.facility.isin(filtered_facilities + [session_state.warehouse])]
    

    session_state.scenario_data = initialize_scenario(session_state, filtered_facilities)
    scenario_name = session_state.scenario_data["Scenario"]
    session_state.scenario_data["Facility_DF"]['type'] = session_state.scenario_data["Facility_DF"].apply(lambda f: 'Warehouse' if f['facility']==session_state.warehouse else f['type'], axis=1)
    session_state.scenario_file = f"filestore/Scenario {scenario_name}.xlsx"
    #st.experimental_rerun()
    return True

def validate_order_data(session_state, uploaded_file):
    """Validate the format of an order data file.
    Args:
        session_state (object): a new SessionState object
        uploaded_file (file): an excel file containing order data
    Returns:
        if uploaded_file meets criteria:
            (bool): True
            (str): a string containing a success message
        
        if uploaded_file does not meets criteria:
            (bool): False           
            error (str): a string containing an error message
            exception (str): a string containing an error message        
    """ 

    try:
        """ Check all columns are found (Date, Facility ID, Facility, Volume) """
        user_file_sample = pd.read_excel(uploaded_file)
        simplified_columns = [" ".join(x.lower().split("_")) for x in user_file_sample.columns]
        order_cols = {"id": "Facility ID", "facility": "Facility", "volume": "Volume (cubic meters)"}
        col_dict = {}
        for col in order_cols.keys():
            """ Check to see if all the words are in at least one column. Remove facility from facility id so facility doesn't return two Trues. """
            lst = [col in x if "id" not in x else col in x.replace("facility", "") for x in simplified_columns]
            res = [i for i, val in enumerate(lst) if val]
            col_dict[res[0]] = col
        sorted_index = np.sort([*col_dict])
        new_columns = [order_cols[col_dict[i]] for i in sorted_index]
        user_file_sample.columns = new_columns

        """ Check all facilities are found in reference data """
        ref_facility = pd.read_excel(session_state.ref_file, sheet_name = "Facility")
        facility_set_diff = set(user_file_sample.Facility) - set(ref_facility.facility)
        assert facility_set_diff == set(), f'{[fac for fac in facility_set_diff]} facilities not found in reference sheet'

        facility_id_set_diff = set(user_file_sample["Facility ID"].apply(lambda c: str(c)[:8])) - set(ref_facility["facility_id"].apply(lambda c: str(c)[:8]))
        assert facility_id_set_diff == set(), f'{[fac for fac in facility_id_set_diff]} facility IDs not found in reference sheet'

        return True, "Upload succesful"
    except AssertionError as error:
        st.error(f"Application Error: {error}\n{traceback.format_exc()}")
        return False, error
    except Exception as exception:
        st.error(f"Application Error: {exception}\n{traceback.format_exc()}")
        return False, exception


def display_raw_data(session_state):
    """Display a map and let user filter out facilities.
    Args:
        session_state (object): a new SessionState object
    """ 

    ref_facility = pd.read_excel(session_state.ref_file, sheet_name = "Facility")  
    ref_facility["facility_id"] = ref_facility["facility_id"].apply(lambda c: str(c)[:8])
    ref_facility["latitude"] = ref_facility["latitude"].astype(float)
    ref_facility["longitude"] = ref_facility["longitude"].astype(float)
    ref_facility['type'] = ref_facility.apply(lambda f: 'Warehouse' if f['facility']==session_state.warehouse else f['type'], axis=1)

    if "Delivery" in pd.ExcelFile(session_state.user_file).sheet_names:
        deliveries = pd.read_excel(session_state.user_file, sheet_name='Delivery')
        order_data = deliveries.groupby(['Dispatch Destination']).agg({'Loading Volume':'sum'}).reset_index()
        order_data.columns = ['facility', 'vol']
        
        dispatch_locations = order_data['facility'].to_list()
        ref_locations = ref_facility['facility'].to_list()
        missed_facilities = [fac for fac in dispatch_locations if fac not in ref_locations]
        if len(missed_facilities) > 0: 
            st.markdown(f"#### Warning: The following dispatch destinations are not recognized. Please review and maintain reference data as-needed: \n{', '.join(missed_facilities)}")

        order_data = ref_facility.merge(order_data, on = "facility", how="left")

    else: 
        order_data = pd.read_excel(session_state.user_file).sort_values("Facility")
        order_data["Facility ID"] = order_data["Facility ID"].apply(lambda c: str(c)[:8])
        order_data = ref_facility.merge(order_data, right_on = "Facility ID", left_on = "facility_id", how="left").rename(columns = {"Volume (cubic meters)" : "vol"})

    with st.expander("Select Facilities", expanded=False):
        facility_stats, space = st.columns([1, 2])
        select_facility_column, space, map_column = st.columns((2.5, 0.5, 5))
    
        order_keys = order_data[order_data["vol"] > 0].facility
        order_data_dict = {key: "" for key in order_keys}
        for k, v in order_data_dict.items():
            vol = order_data[order_data.facility == k]["vol"].iloc[0]
            order_data_dict[k] = select_facility_column.checkbox(f"{str(k)} ({vol :.2f} cubic meters)", value=k, key = str(k))
        
        filtered_facilities = list(itertools.compress(order_data_dict.keys(), order_data_dict.values()))
        order_data = order_data[order_data.facility.isin(filtered_facilities + [session_state.warehouse])]
        facility_stats.markdown(f"#### Total # of Facilities: {len(order_data) - 1}")
        facility_stats.markdown(f"#### Total Volume Demand: {order_data['vol'].sum():.2f} m<sup>3</sup>", unsafe_allow_html=True)
        facility_stats.markdown("***")

        m = map_orders(session_state, order_data)
        with map_column.container():
            folium_static(m, width = 900, height = 600)

    initialize_button = st.button("Initialize scenario")
    if initialize_button:
        session_state.scenario_data = initialize_scenario(session_state, filtered_facilities)
        scenario_name = session_state.scenario_data["Scenario"]
        session_state.scenario_data["Facility_DF"]['type'] = session_state.scenario_data["Facility_DF"].apply(lambda f: 'Warehouse' if f['facility']==session_state.warehouse else f['type'], axis=1)
        session_state.scenario_file = f"filestore/Scenario {scenario_name}.xlsx"
        st.experimental_rerun()

def display_scenario_data(session_state):
    """Display a map and let user view scenario data.
    Args:
        session_state (object): a new SessionState object
    """ 

    order_data = session_state.scenario_data["Facility_DF"].copy()
    st.markdown(f"### Origin: {session_state.warehouse}")
    with st.expander("Facilities", expanded=False):
        facility_column, space, map_column  = st.columns((2.5, 0.5, 5))
        facility_column.markdown(f"#### Total # of Facilities: {len(order_data) - 1}")
        facility_column.markdown(f"#### Total Volume Demand: {order_data.vol.sum():.2f} m<sup>3</sup>", unsafe_allow_html=True)
        facility_column.markdown("***")

        m = map_orders(session_state, order_data)

        for index, row in order_data.iterrows():
            if row['type'] != 'Warehouse': 
                facility_column.markdown(f"{row['facility']} ({row['vol']:.2f} m<sup>3</sup>)", unsafe_allow_html=True)

        with map_column.container():
            folium_static(m, width = 900, height = 600)
    
def map_orders(session_state, order_data):
    """Display a map and let user view facility locations.
    Args:
        session_state (object): a new SessionState object
        order_data (pandas.core.frame.DataFrame): a pandas dataframe object containing order data
        
    Returns:
        m (folium.folium.Map): a folium map object
    """ 

    """ Duplicate the facility name column before setting it as an index""" 
    order_data["name"] = order_data["facility"]

    map_scale = 6 if "Zambia DRT Lusaka Reg" in session_state.ref_file else 10
    """ Center the map around the warehouse""" 
    m = folium.Map(location = [order_data[order_data.type == "Warehouse"].latitude, order_data[order_data.type == "Warehouse"].longitude], zoom_start = map_scale)
    order_data = order_data.set_index("facility")
    folium.Marker([order_data[order_data.type == "Warehouse"].latitude, order_data[order_data.type == "Warehouse"].longitude], tooltip = order_data[order_data.type == "Warehouse"].name.iloc[0]).add_to(m)
    """ Plot the rest of the facility points"""
    order_data = order_data[order_data.type != "Warehouse"]
    for i in range(0, len(order_data)):
        facility = order_data.iloc[i].name
        facility_coord = [order_data.loc[facility, "longitude"], order_data.loc[facility, "latitude"]]

        folium.Marker([facility_coord[1], facility_coord[0]],
                    popup = folium.Popup(facility + "<br>Volume requested: " + str(order_data.iloc[i].vol) + " m<sup>3</sup>", max_width=350, min_width=250),
                    tooltip = facility,
                    icon=folium.Icon(color='darkgreen', icon_color='darkgreen')
                    ).add_to(m)
    return m
