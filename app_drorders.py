"""This module contains functions used to analyze order data."""

import streamlit as st
import pandas as pd
import os
from math import ceil
import traceback
from zipfile import ZipFile
import base64
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import PatternFill, colors, Font, Alignment
import altair as alt
from utils import get_binary_file_downloader_html
from datetime import datetime

import random
import string

from ortools.linear_solver import pywraplp

ORDER_STATUS = ['New', 'Planned', 'Released', 'Picking', 'Picked', 'Packed', 'Staged', 'Loading', 'Shipped', 'Cancelled']

def load_order_details(fname):
    """Load and transform a dataframe containing order data.
    Args:
        fname (file): a excel file containing order data
    
    Returns:
        order_details_df (pandas.core.frame.DataFrame): a dataframe object containing order data
    """ 

    order_details_df = pd.read_excel(fname) 
    order_details_df['Customer ID'] = order_details_df['Customer ID'].apply(lambda c: str(c)[:8])
    return order_details_df

def load_facility_mapping(fname):
    """Load and transform dataframe containing FAC mapping data.
    Args:
        fname (file): a excel file containing order data
    
    Returns:
        fac_df (pandas.core.frame.DataFrame): a dataframe object containing FAC mapping data
    """    

    fac_df = pd.read_excel(fname, sheet_name='Fac Mapping')
    fac_df['Customer ID'] = fac_df['WMS_Fac_Code'].apply(lambda c: str(c)[:8])
    fac_df['Hub_Code'] = fac_df['Hub_Code'].apply(lambda c: str(c)[:8])
    fac_df['Delivery_Destination_Code'] = fac_df['Delivery_Destination_Code'].apply(lambda c: str(c)[:8])
    fac_df['District_Health_Office_Code'] = fac_df['District_Health_Office_Code'].apply(lambda c: str(c)[:8])
    return fac_df 

def load_orders(fname, tabs, allowed_status): 
    """Load order data.
    Args:
        fname (file): a excel file containing order data
        
        tabs (list): a list object containing sheet names of an excel file
        
        allowed_status (list): a list object containing permitted orderers     
    
    Returns:        
         orders_df (pandas.core.frame.DataFrame): a dataframe object containing order data    
    """ 

    wb = load_workbook(filename = fname, data_only=True)

    valid_order_prefix = {'AR', 'CO', 'DI', 'EM', 'EQ', 'GL', 'HI', 'IR', 'IT', 'KI', 'LA', 'LE', 'MA', 
                            'MC', 'MD', 'MS', 'NP', 'NT', 'QC', 'RE', 'RH', 'RN', 'TB', 'TP', 'WE', 'ZN',}

    orders = []
    for name in wb.sheetnames if tabs is None else tabs:
        if name in wb.sheetnames: 
            df = pd.DataFrame(wb[name].values)
            if df.shape[0] == 0: 
                continue

            st_order_cnt = len(orders)
            if pd.notna(df.iloc[0, 0]) and df.iloc[0, 0].upper() in {"ROUTE", "ORDER ID"}: 
                df.columns = [c.upper() for c in df.iloc[0]]
                df = df.iloc[1:, :]
                for _, row in df.iterrows(): 
                    if pd.notna(row['ORDER ID']) and row['ORDER ID'].strip()[:2].isalpha(): 
                        orders.append([row['ROUTE'], row['ORDER ID']])
            else: 
                try: 
                    orders_ignored = 0
                    order_line_start = False
                    for _, row in df.iterrows(): 
                        if (pd.notna(row[1]) and type(row[1])==str and row[1].strip().upper()=='ROUTE') or (pd.notna(row[0]) and type(row[0])==str and row[0].strip().upper()=='ROUTE'): 
                            order_line_start = True
                        elif order_line_start and pd.notna(row[1]): 
                            route = row[1].strip()
                            for c in range(3, len(row)): 
                                if pd.notna(row[c]) and type(row[c]) == str and len(row[c]) >= 9 and row[c].strip()[:2] in valid_order_prefix and row[c][5:8].isdigit(): 
                                    if allowed_status is None or (pd.notna(row[c+1]) and row[c+1].strip() in allowed_status): 
                                        orders.append([route, row[c].strip()])
                                    else:
                                        orders_ignored += 1
                    if orders_ignored > 0: 
                        st.markdown(f"##### Note: {orders_ignored} orders were ignored from {name} due to status not in allowed list.")
                except Exception as e:
                    st.error(f"Application Error: {e}\n{traceback.format_exc()}")
                    st.error(f"Please verify file format")
            
            print(f"Read {len(orders)-st_order_cnt} orders from sheet {name}")

    orders_df = pd.DataFrame(orders, columns=['Route', 'Customer Order Number'])
    return orders_df

def est_loading_volume(v): 
    """Estimate loading volume for a delivery.
    Args:
        v (float): a float value denoting loading volume
    
    Returns:
         (float): a float value denoting loading volume        
    """      

    small_box = 0.011024    # (26.5 cm x 26 cm x 16 cm)
    medium_box = 0.0296     # (34 cm x 33.5 cm x 26 cm)

    return max(medium_box, v*1.25)

    # if v >= 0.8:
    #     return max(1, v)*1.5
        
    # v *= 1.25    # 120% 
    # if v <= small_box: 
    #     return small_box

    # medium_box_cnt = v//medium_box
    # remaining_vol = v - medium_box_cnt*medium_box
    # if remaining_vol <= small_box: 
    #     return medium_box_cnt*medium_box + small_box
    # else:
    #     return (medium_box_cnt+1)*medium_box

@st.cache(suppress_st_warning=True, show_spinner=False, allow_output_mutation=True)
def estimate_order_volume(order_details_df):    
    """Estimate order volume.
    Args:
        order_details_df (pandas.core.frame.DataFrame) a dataframe object containing order data
  
    Returns:     
         order_vol_df (pandas.core.frame.DataFrame): a dataframe object containing order data   
    """      
 
    order_vol_df = order_details_df.groupby(['Customer Order Number', 'Customer ID']).agg({'Line Weight':'sum', 'Line Volume':'sum', 
                        'Impute Flag':'max', 'Review Flag':'max'}).reset_index()
    order_vol_df['Loading Weight'] = order_vol_df['Line Weight']*1.2
    order_vol_df['Loading Volume'] = order_vol_df['Line Volume'].apply(est_loading_volume)
    order_vol_df = order_vol_df[['Customer Order Number', 'Customer ID', 'Line Weight', 'Line Volume', 'Loading Weight', 'Loading Volume', 'Impute Flag', 'Review Flag']]
    return order_vol_df

@st.cache(suppress_st_warning=True, show_spinner=False, allow_output_mutation=True)
def get_sheet_names(fname): 
    """Get sheet names of an excel file.
    Args:
        fname (file): an excel file 
  
    Returns:    
         (list): a list object containing sheet names of an excel file   
    """  

    return pd.ExcelFile(fname).sheet_names

def evaluate_orders(orders_df, order_details_df, fac_df):
    """Allow user to evaluate order data.
    Args:
        orders_df (pandas.core.frame.DataFrame): a dataframe object containing order data           
        
        order_details_df (pandas.core.frame.DataFrame) a dataframe object containing order data
        
        fac_df (pandas.core.frame.DataFrame): a dataframe object containing FAC mapping data        
  
    Returns: 
        dro_delivery_df (pandas.core.frame.DataFrame): a dataframe object containing DRO delivery data           
        
        dro_info_df (pandas.core.frame.DataFrame) a dataframe object containing DRO data
        
        dro_details_df (pandas.core.frame.DataFrame): a dataframe object containing order data
    """

    order_vol_df = estimate_order_volume(order_details_df)
    dro_info_df = pd.merge(orders_df, order_vol_df, on='Customer Order Number', how='left')
    
    n_orders = len(orders_df)
    n_orders_matched = dro_info_df['Customer ID'].count()
    if n_orders != n_orders_matched: 
        st.warning(f"Out of {n_orders} orders detected, missing order details for {n_orders - n_orders_matched}")
        if n_orders_matched < 0.9*n_orders: 
            st.warning(f"Please notify PSM team to ensure the latest order details are updated.")
            

    dro_info_df = pd.merge(dro_info_df, fac_df[['Customer ID', 'Customer_Name', 'District', 'Province']], on='Customer ID', how='left')
    dro_delivery_df = dro_info_df.groupby(['Route', 'Customer_Name', 'Customer ID']).agg({'Loading Weight':'sum', 'Loading Volume':'sum', 'Impute Flag':'max', 'Review Flag':'max'}).reset_index()
    
    dro_info_df = dro_info_df[['Route', 'Customer Order Number', 'Customer_Name', 'Customer ID', 'District', 'Province', 'Line Weight', 'Line Volume', 'Loading Weight', 'Loading Volume', 'Impute Flag', 'Review Flag']]
    dro_info_df.columns=['Route', 'Order Number', 'Customer Name', 'Customer ID', 'District', 'Province', 'Line Weight', 'Line Volume', 'Loading Weight', 'Loading Volume', 'Impute Flag', 'Review Flag']


    dro_delivery_df = pd.merge(dro_delivery_df, fac_df[['Customer ID', 'District', 'Province', 'Hub_Name', 'Hub_Code', 'District_Health_Office', 'District_Health_Office_Code',
                                                        'Delivery_Destination_Name', 'Delivery_Destination_Code']], 
                                on='Customer ID', how='left')

    dro_delivery_df['Dispatch Destination'] = dro_delivery_df.apply(lambda r: r['Delivery_Destination_Name'] if r['Hub_Code'] == '999999' else r['Hub_Name'], axis=1)
    dro_delivery_df['Dispatch Destination Code'] = dro_delivery_df.apply(lambda r: r['Delivery_Destination_Code'] if r['Hub_Code'] == '999999' else r['Hub_Code'], axis=1)

    dro_delivery_df = dro_delivery_df[['Route', 'Customer_Name', 'Hub_Name', 'District_Health_Office', 'Customer ID', 'Hub_Code', 'District_Health_Office_Code', 'District', 'Province', 
                                    'Loading Weight', 'Loading Volume', 'Impute Flag', 'Review Flag', 'Dispatch Destination', 'Dispatch Destination Code']]
    dro_delivery_df.columns=['Route', 'Customer Name', 'Hub Name', 'District Health_Office', 'Customer ID', 'Hub Code', 'District Health Office Code', 'District', 'Province',
                                    'Loading Weight', 'Loading Volume', 'Impute Flag', 'Review Flag', 'Dispatch Destination', 'Dispatch Destination Code']
    dro_delivery_df['Dedicated Trucks'] = dro_delivery_df.apply(lambda r: 'Yes' if r['Dispatch Destination Code'][:4] == '9999' else 'No', axis=1)

    dro_details_df = pd.merge(order_details_df, orders_df['Customer Order Number'], on='Customer Order Number')
 
    return dro_delivery_df, dro_info_df, dro_details_df

def write_df_to_excel(wb, sheetname, df, fontsize=10): 
    """Write a dataframe to excel.
    Args:
        wb (file): an excel file 
        
        sheetname (str): a sheet name of an excel file
        
        df (pandas.core.frame.DataFrame): a dataframe object         
        
        fontsize (int, optional): a integer value denoting a font size (default is 10)
  
    """    

    ws = wb[sheetname]
    rows = dataframe_to_rows(df, index=False, header=False)
    for r_idx, row in enumerate(rows, 2):
        for c_idx, value in enumerate(row, 1):
            try:
                ws.cell(row=r_idx, column=c_idx).value = value
            except:
                ws.cell(row=r_idx, column=c_idx).value = None

def app_drorders(session_state):
    """Allow user to evaluate orders.
    Args:
        session_state (object): a new SessionState object 
  
    """      

    st.markdown("***")
    dro_delivery_df = None
    dro_info_df = None
    dro_details_df = None

    st.markdown(f"###### Note: Order details last updated on {datetime.fromtimestamp(os.path.getmtime(session_state.dro_order_details_file)).strftime('%m/%d/%Y %H:%M:%S')}")

    left_m, col_browse, mid_m, col_button, right_m = st.columns([0.2, 5, 0.2, 3, 0.2])
    order_file = col_browse.file_uploader("Route Status File", type=["xlsx"], key="RouteStatus")
    if order_file is not None:
        if col_button.button("Upload Route Status File"): 
            dro_info_df = None
            dro_details_df = None
            fname = f"A{session_state.dro_scenario_key}_{order_file.name}"
            with open(os.path.join("./filestore", fname), 'wb') as f:
                f.write(order_file.getbuffer())
                session_state.dro_order_file = os.path.join("./filestore", fname)
    else: 
        session_state.dro_order_file = ''

    st.markdown("***")
    if session_state.dro_order_file != '': 

        left_m, sheet_sel, mid_m, status_sel, right_m = st.columns([0.2, 5, 0.2, 3, 0.2])

        sheet_names = get_sheet_names(session_state.dro_order_file)
        tabs = sheet_sel.multiselect("Select ordres in the following tabs", sheet_names,  sheet_names)

        allowed_status = None
        with status_sel: 
            if not st.checkbox("Include orders with all statuses", value=True): 
                allowed_status = st.multiselect("Include only the following status", ORDER_STATUS, default="New")

        st.markdown("***")
        if len(tabs) > 0 and st.button("Evaluate Orders from Selected Routes"): 

            oe_progress_bar = st.progress(0)

            with st.spinner('Loading target orders ...'):
                orders_df = load_orders(session_state.dro_order_file, tabs, allowed_status)
                if len(orders_df) == 0: 
                    st.warning("No orders were identified. Please check your file.")
                    st.stop()
            oe_progress_bar.progress(10)

            with st.spinner('Loading order details ...'):
                order_details_df = load_order_details(session_state.dro_order_details_file)
            oe_progress_bar.progress(30)

            with st.spinner('Loading network structure ...'):
                fac_df = load_facility_mapping(session_state.dro_facility_mapping_file)
            oe_progress_bar.progress(60)
            
            with st.spinner('Evaluate orders ...'):
                dro_delivery_df, dro_info_df, dro_details_df = evaluate_orders(orders_df, order_details_df, fac_df)
            oe_progress_bar.progress(80)

            with st.spinner('Generating order evaluation report ...'):
                dro_f = f"./filestore/Zambia DRO Order Evaluation {session_state.dro_scenario_key}.xlsm"            
                wb = load_workbook(session_state.dro_order_evaluation_template_file, keep_vba=True)
                write_df_to_excel(wb, "Delivery", dro_delivery_df)
                write_df_to_excel(wb, "Order Info", dro_info_df)
                write_df_to_excel(wb, "Order Details", dro_details_df)
                write_df_to_excel(wb, "Facility Ref", fac_df[['Customer_Name', 'WMS_Fac_Code']])
                write_df_to_excel(wb, "Delivery Original", dro_delivery_df)
                write_df_to_excel(wb, "Order Info Original", dro_info_df)
                write_df_to_excel(wb, "Order Details Original", dro_details_df)
                wb.save(dro_f)
            oe_progress_bar.progress(90)

            with st.spinner('Generating download link ...'):
                st.markdown(get_binary_file_downloader_html(dro_f, 'DRO Order Evaluation Result'), unsafe_allow_html=True)
            oe_progress_bar.progress(100)

        if dro_delivery_df is not None:     # visualization
            vol_summary_df = dro_delivery_df.groupby(['Route', 'Dispatch Destination']).agg({'Loading Volume':'sum'}).reset_index()
            vol_summary_df.sort_values(['Route', 'Loading Volume'], ascending = [True, False], inplace = True, na_position ='last')

            nsteps = len(vol_summary_df)
            step_width = min(100, int(1000/max(1, nsteps)))

            fig1 = alt.Chart(vol_summary_df).mark_bar(
                        cornerRadiusTopLeft=3,
                        cornerRadiusTopRight=3
                    ).encode( 
                        alt.Color('Route:N'),
                        x=alt.X(f'Dispatch Destination:N', 
                                    axis=alt.Axis(labelFontSize=11, titleFontSize=16, labelAngle=-30), 
                                    sort=alt.EncodingSortField(field="oading Volume", order='descending')),
                        y=alt.Y('Loading Volume:Q', axis=alt.Axis(labelFontSize=12, titleFontSize=16, title="Loading Volume (m3)")),
                        tooltip=['Route', 'Dispatch Destination', 'Loading Volume']
                    ).properties(height=400, width=alt.Step(step_width), title=f"Order Evaluation - Loading Volume by Dispatch Destination").configure_title(fontSize=18)
            st.altair_chart(fig1)
                           

if __name__ == "__main__":

    import SessionState

    st.set_page_config(layout="wide", initial_sidebar_state='auto')
    st.title("Zambia Dynamic Routing for Orders (DRO)")

    session_state = SessionState.get(dro_order_details_file = r"./data/Zambia DRO Order Details.xlsx", 
                                     dro_facility_mapping_file = r"./data/Zambia ALL Facility Mapping.xlsx",
                                     dro_order_evaluation_template_file = r"./data/Zambia DRO Order Evaluation Template.xlsm", 
                                     dro_scenario_key = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6)), 
                                     dro_order_file ="")
    app_drorders(session_state)
